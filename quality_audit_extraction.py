#!/usr/bin/env python3
"""
qa_excel_import.py
==================
Imports one or more Quality Audit Excel files into MongoDB (Quality_Audit.Audit_Data).

Red-text detection (for Misses / Serious field):
  • Case 1 – Entire cell has a red font style  → whole evidence text = misses
  • Case 2 – Cell contains rich-text runs       → only the red-coloured runs = misses

Usage
-----
  # Single file
  python qa_excel_import.py "Quality_Audit_HDFC.xlsx"

  # All xlsx in a folder
  python qa_excel_import.py folder/

  # Dry-run (print what would be inserted, no DB write)
  python qa_excel_import.py "Quality_Audit_HDFC.xlsx" --dry-run

  # Skip already-imported files (match by project_id + audit_date)
  python qa_excel_import.py folder/ --skip-existing

Environment / Config
--------------------
  Set MONGO_URI in a .env file or as an environment variable.
  Default: mongodb://localhost:27017
"""

import os
import sys
import glob
import json
import argparse
import zipfile
from datetime import datetime
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from pymongo import MongoClient

load_dotenv()

# ─── MongoDB ──────────────────────────────────────────────────────────────────
MONGO_URI   = os.getenv("MONGO_URI", "mongodb://localhost:27017")
DB_NAME     = "Quality_Audit"
COLLECTION  = "Audit_Data"

# ─── Section boundary map  (row label → section key) ─────────────────────────
SECTION_MARKERS = {
    "(a) audit planning":   "section_a",
    "(b) audit execution":  "section_b",
    "(b) audit execution:": "section_b",
    "(c) audit reporting":  "section_c",
    "(c) audit reporting:": "section_c",
    "(d) quality control":  "section_d",
    "(d) quality control:": "section_d",
    "(e) optimization":     "section_e",
    "(e) optimization:":    "section_e",
}

STOP_ROWS = {"total", "scaled total", "totals"}

# Red RGB values we recognise (openpyxl and raw XML)
RED_RGBS = {"FFFF0000", "FF0000", "FF3333", "CC0000", "FF1111",
            "FFCC0000", "FFB22222", "FFDC143C", "FFFF0000"}


# ─── Helper: is this RGB red? ─────────────────────────────────────────────────
def _is_red_rgb(rgb: str | None) -> bool:
    if not rgb:
        return False
    return rgb.upper() in RED_RGBS or rgb.upper().endswith("0000")  # e.g. FF??0000 patterns


# ─── Core extractor ───────────────────────────────────────────────────────────

class ExcelQAParser:
    """Parses a single Quality Audit Excel file."""

    NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    def __init__(self, path: str):
        self.path = path
        self._wb        = None   # openpyxl workbook (data_only)
        self._ws        = None   # Checklist sheet
        self._fonts     = []     # parsed style fonts
        self._ss        = []     # shared strings: {'type':'plain'|'rich', ...}
        self._cell_reds = {}     # {coordinate: bool}  whole-cell red
        self._cell_misses = {}   # {coordinate: str}   red text extracted
        self._load()

    # ── Load & pre-process ────────────────────────────────────────────────────

    def _load(self):
        self._wb = load_workbook(self.path, data_only=True)
        self._ws = self._wb["Checklist"]
        self._parse_xml()

    def _parse_xml(self):
        """
        Read styles + sharedStrings from raw XML to get run-level colour data
        that openpyxl's high-level API doesn't expose.
        """
        with zipfile.ZipFile(self.path, "r") as z:
            styles_xml = z.read("xl/styles.xml")
            ss_xml     = z.read("xl/sharedStrings.xml")
            sheet_xml  = z.read("xl/worksheets/sheet1.xml")

        ns = self.NS

        # ── 1. Parse font list from styles ────────────────────────────────────
        styles_tree = etree.fromstring(styles_xml)
        self._fonts = []
        for font in styles_tree.findall(".//a:fonts/a:font", ns):
            color_el = font.find("a:color", ns)
            is_red   = False
            if color_el is not None:
                rgb = color_el.get("rgb")
                if _is_red_rgb(rgb):
                    is_red = True
            self._fonts.append(is_red)

        # Build xf → fontId map
        xfs = styles_tree.findall(".//a:cellXfs/a:xf", ns)
        self._xf_fonts = [int(xf.get("fontId", 0)) for xf in xfs]

        # ── 2. Parse shared strings ───────────────────────────────────────────
        ss_tree = etree.fromstring(ss_xml)
        self._ss = []
        for si in ss_tree.findall("a:si", ns):
            runs = si.findall("a:r", ns)
            if runs:
                parts = []
                for r in runs:
                    rpr = r.find("a:rPr", ns)
                    t   = r.find("a:t",   ns)
                    text = (t.text or "") if t is not None else ""
                    is_red = False
                    if rpr is not None:
                        c_el = rpr.find("a:color", ns)
                        if c_el is not None and _is_red_rgb(c_el.get("rgb")):
                            is_red = True
                    parts.append({"text": text, "is_red": is_red})
                self._ss.append({"type": "rich", "parts": parts})
            else:
                all_t = si.findall(".//a:t", ns)
                self._ss.append({"type": "plain", "text": "".join(t.text or "" for t in all_t)})

        # ── 3. For every cell in sheet col C: determine whole-cell red ────────
        sheet_tree = etree.fromstring(sheet_xml)
        self._whole_cell_red = {}  # {(row_int, col_int): True}
        for row_el in sheet_tree.findall(".//a:row", ns):
            for c in row_el.findall("a:c", ns):
                ref   = c.get("r", "")
                s_idx = c.get("s")
                if s_idx is None:
                    continue
                s_int = int(s_idx)
                if s_int < len(self._xf_fonts):
                    fid = self._xf_fonts[s_int]
                    if fid < len(self._fonts) and self._fonts[fid]:
                        # Parse ref like C28 → col 3, row 28
                        col_str = "".join(ch for ch in ref if ch.isalpha())
                        row_str = "".join(ch for ch in ref if ch.isdigit())
                        if col_str and row_str:
                            col_num = sum(
                                (ord(ch) - ord("A") + 1) * (26 ** i)
                                for i, ch in enumerate(reversed(col_str.upper()))
                            )
                            self._whole_cell_red[(int(row_str), col_num)] = True

    # ── Public parser ─────────────────────────────────────────────────────────

    def parse(self) -> dict:
        ws = self._ws
        doc = {}

        # ── Header fields (rows 1–9) ──────────────────────────────────────────
        doc["project_id"]   = self._v(ws, 1, 2)   # B1
        doc["project_name"] = self._v(ws, 2, 2)   # B2
        doc["client_name"]  = self._v(ws, 3, 2)   # B3
        doc["type_of_audit"]= self._v(ws, 4, 2)   # B4

        # Audit period — C5 = from, E5 = to
        doc["audit_period_from"] = self._fmt_date(self._v(ws, 5, 3))
        doc["audit_period_to"]   = self._fmt_date(self._v(ws, 5, 5))

        # Auditor — C6 = name, emp ID via HR Ref lookup
        auditor_name = self._v(ws, 6, 3)   # C6
        doc["audit_given_by_name"]   = auditor_name
        doc["audit_given_by_emp_id"] = self._lookup_emp_id(auditor_name)

        doc["project_tl"]  = self._v(ws, 7, 2)   # B7
        doc["project_pnd"] = self._v(ws, 8, 2)   # B8
        doc["audit_date"]  = self._v(ws, 9, 2)   # B9

        # ── Sections ─────────────────────────────────────────────────────────
        sections = {k: [] for k in ("section_a","section_b","section_c","section_d","section_e")}
        current_section = None

        for row_num in range(10, ws.max_row + 1):
            a_val = self._v(ws, row_num, 1)   # Particular
            b_val = self._v(ws, row_num, 2)   # YES/NO/NA
            c_val = self._cell_text(ws, row_num, 3)  # Evidence (full text)
            d_val = self._v(ws, row_num, 4)   # Score
            e_val = self._v(ws, row_num, 5)   # Ideal Score

            if not a_val:
                continue

            a_lower = str(a_val).strip().lower()

            # Stop at TOTAL row
            if a_lower in STOP_ROWS or a_lower.startswith("total") or a_lower.startswith("scaled"):
                break

            # Section marker?
            if a_lower in SECTION_MARKERS:
                current_section = SECTION_MARKERS[a_lower]
                continue

            # Skip header row
            if a_lower in ("particulars", "particular"):
                continue

            if current_section is None:
                continue

            # ── Extract misses from red text ──────────────────────────────────
            evidence, misses = self._extract_evidence_and_misses(row_num, c_val)

            # ── Parse score safely ────────────────────────────────────────────
            score       = self._safe_num(d_val)
            ideal_score = self._safe_num(e_val)

            sections[current_section].append({
                "particular":  str(a_val).strip(),
                "response":    str(b_val).strip().upper() if b_val else "",
                "evidence":    evidence,
                "misses":      misses,
                "score":       score,
                "ideal_score": ideal_score,
            })

        doc.update(sections)

        # ── Totals from TOTAL row ─────────────────────────────────────────────
        for row_num in range(10, ws.max_row + 1):
            a_val = self._v(ws, row_num, 1)
            if a_val and str(a_val).strip().lower() == "total":
                doc["total_score"] = self._safe_num(self._v(ws, row_num, 4))
                doc["ideal_score"] = self._safe_num(self._v(ws, row_num, 5))
                break

        # Compute scaled total
        t  = doc.get("total_score", 0) or 0
        id = doc.get("ideal_score", 0) or 0
        doc["scaled_total"] = round(t * 100 / id, 2) if id > 0 else 0

        # ── Improvements (not in old-format Excel — set empty) ────────────────
        doc["improvements"] = ""

        # ── Metadata ─────────────────────────────────────────────────────────
        doc["status"]          = "submitted"
        doc["submitted_by"]    = "imported"
        doc["submitted_email"] = "imported"
        doc["submitted_at"]    = datetime.utcnow()
        doc["source_file"]     = Path(self.path).name

        return doc

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _v(self, ws, row: int, col: int):
        """Get plain cell value."""
        cell = ws.cell(row=row, column=col)
        val  = cell.value
        if val is None:
            return ""
        # Strip formula fallback
        if isinstance(val, str) and val.startswith("="):
            return ""
        return val

    def _cell_text(self, ws, row: int, col: int) -> str:
        """Return full text of cell (rich or plain)."""
        val = self._v(ws, row, col)
        if val is None:
            return ""
        return str(val).strip()

    def _extract_evidence_and_misses(self, row_num: int, full_text: str) -> tuple[str, str]:
        """
        Returns (evidence_text, misses_text).

        Strategy:
          1. Check shared string index for this cell → if rich text with red runs
             → separate red runs as misses, non-red as evidence
          2. Else check if whole-cell font is red → entire text is both evidence & misses
          3. Else → misses = ""
        """
        col_num = 3   # column C

        # Try to get shared string index from raw XML (already parsed in _parse_xml)
        # We need to find the <c> value for this cell — easier via openpyxl
        ws   = self._ws
        cell = ws.cell(row=row_num, column=col_num)

        if not full_text:
            return "", ""

        # ── Case 1: whole-cell red font ───────────────────────────────────────
        if self._whole_cell_red.get((row_num, col_num)):
            return full_text, full_text   # evidence = full, misses = full

        # ── Case 2: rich text with red runs ───────────────────────────────────
        # openpyxl stores the shared string index in cell._value when t='s'
        # We need to know if the cell uses a shared string
        # openpyxl doesn't directly expose the SS index on a read cell
        # WORKAROUND: match full_text against our parsed shared strings
        misses_parts = []
        normal_parts = []
        found_rich   = False

        for ss in self._ss:
            if ss["type"] != "rich":
                continue
            ss_full = "".join(p["text"] for p in ss["parts"])
            # Normalise whitespace for matching
            if ss_full.strip() == full_text.strip():
                found_rich = True
                for p in ss["parts"]:
                    if p["is_red"]:
                        misses_parts.append(p["text"])
                    else:
                        normal_parts.append(p["text"])
                break

        if found_rich:
            misses_text  = "".join(misses_parts).strip()
            # Evidence = full text (preserve all context)
            return full_text, misses_text

        # ── Case 3: plain text, no red ────────────────────────────────────────
        return full_text, ""

    def _safe_num(self, val) -> float:
        """Convert a value to float, return 0 on failure."""
        if val is None or val == "":
            return 0
        try:
            return float(str(val).replace(",", "").strip())
        except (ValueError, TypeError):
            return 0

    def _fmt_date(self, val) -> str:
        """Return ISO date string or original string."""
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        if val is None or val == "":
            return ""
        s = str(val).strip()
        # Try to parse common formats
        for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s.split(" ")[0], fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        return s

    def _lookup_emp_id(self, name: str) -> str:
        """Look up employee ID from HR Ref sheet by name."""
        if not name or "HR Ref" not in self._wb.sheetnames:
            return ""
        ws_hr = self._wb["HR Ref"]
        name_upper = str(name).strip().upper()
        for row in ws_hr.iter_rows(min_row=2, values_only=True):
            # Employee name in col E (index 4), JHS ID in col F (index 5)
            if len(row) >= 6 and row[4] and str(row[4]).strip().upper() == name_upper:
                return str(row[5]).strip() if row[5] else ""
        return ""


# ─── DB writer ────────────────────────────────────────────────────────────────

def insert_to_mongo(doc: dict, dry_run: bool, skip_existing: bool) -> str:
    """
    Insert doc into MongoDB. Returns 'inserted', 'skipped', or 'dry_run'.
    """
    if dry_run:
        return "dry_run"

    client = MongoClient(MONGO_URI)
    coll   = client[DB_NAME][COLLECTION]

    if skip_existing:
        exists = coll.find_one({
            "project_id":  doc.get("project_id"),
            "audit_date":  doc.get("audit_date"),
            "project_tl":  doc.get("project_tl"),
            "client_name": doc.get("client_name"),
        })
        if exists:
            client.close()
            return "skipped"

    coll.insert_one(doc)
    client.close()
    return "inserted"


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import Quality Audit Excel → MongoDB")
    parser.add_argument("--dry-run",       action="store_true", help="Parse only, don't insert")
    parser.add_argument("--skip-existing", action="store_true", help="Skip if project_id+audit_date already exists")
    parser.add_argument("--show-json",     action="store_true", help="Print parsed JSON for each file")
    args = parser.parse_args()

    folder_path = r"C:\Users\vasu.gadde\Downloads\Outlook_Excel_Files"
    # Collect files
    p = Path(folder_path)
    if p.is_dir():
        files = sorted(p.glob("**/*.xlsx"))
    elif p.is_file():
        files = [p]
    else:
        print(f"ERROR: '{folder_path}' is not a file or folder.")
        sys.exit(1)

    if not files:
        print("No .xlsx files found.")
        sys.exit(0)

    print(f"\n{'='*60}")
    print(f"Quality Audit Importer")
    print(f"Files found : {len(files)}")
    print(f"Dry run     : {args.dry_run}")
    print(f"Skip existing: {args.skip_existing}")
    print(f"MongoDB     : {MONGO_URI} → {DB_NAME}.{COLLECTION}")
    print(f"{'='*60}\n")

    results = {"inserted": 0, "skipped": 0, "dry_run": 0, "error": 0}

    for xlsx_path in files:
        print(f"► Processing: {xlsx_path.name}")
        try:
            parser_obj = ExcelQAParser(str(xlsx_path))
            doc = parser_obj.parse()

            # Print summary
            total_items = sum(len(doc.get(f"section_{s}", [])) for s in "abcde")
            items_with_misses = sum(
                1
                for s in "abcde"
                for it in doc.get(f"section_{s}", [])
                if it.get("misses")
            )
            print(f"  Client     : {doc.get('client_name')}")
            print(f"  TL         : {doc.get('project_tl')}")
            print(f"  Audit date : {doc.get('audit_date')}")
            print(f"  Score      : {doc.get('total_score')} / {doc.get('ideal_score')}  (scaled: {doc.get('scaled_total')})")
            print(f"  Sections   : {total_items} items,  {items_with_misses} with Misses/Serious")

            # Show misses detail
            if items_with_misses:
                print("  Misses detected:")
                for s in "abcde":
                    for it in doc.get(f"section_{s}", []):
                        if it.get("misses"):
                            print(f"    [{s.upper()}] {it['particular'][:55]}…")
                            print(f"         Evidence: {it['evidence'][:80]}")
                            print(f"         Misses  : {it['misses'][:80]}")

            if args.show_json:
                # Pretty print without _id
                d2 = {k: v for k, v in doc.items() if k != "_id"}
                # Convert datetime for JSON
                if isinstance(d2.get("submitted_at"), datetime):
                    d2["submitted_at"] = d2["submitted_at"].isoformat()
                print(json.dumps(d2, indent=2, default=str))

            status = insert_to_mongo(doc, args.dry_run, args.skip_existing)
            results[status] += 1
            print(f"  Status     : {status.upper()}\n")

        except Exception as e:
            import traceback
            print(f"  ERROR: {e}")
            traceback.print_exc()
            results["error"] += 1
            print()

    print(f"{'='*60}")
    print(f"Summary: Inserted={results['inserted']}  Skipped={results['skipped']}  "
          f"DryRun={results['dry_run']}  Errors={results['error']}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()