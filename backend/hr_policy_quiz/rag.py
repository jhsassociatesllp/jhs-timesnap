"""
Lightweight RAG pipeline.

Why TF-IDF instead of a neural embedding model:
- No GPU/heavy download needed, ingestion is fast and cheap.
- The document is read ONCE at upload time, chunked, and a bank of
  questions is generated once and cached in MongoDB. Every quiz after
  that is just a random sample from the cached pool - so the LLM is
  never called again for the same document. This is the "cost saving"
  the RAG setup is for: one ingestion cost, unlimited quizzes.
"""
import json
import random
import re
import uuid
from io import BytesIO
from typing import List

import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from backend.hr_policy_quiz.config import settings

try:
    from openai import OpenAI
    _client = OpenAI(api_key=settings.OPENAI_API_KEY) if settings.OPENAI_API_KEY else None
except Exception:
    _client = None


def extract_text(file_bytes: bytes, filename: str) -> str:
    name = filename.lower()
    if name.endswith(".pdf"):
        from pypdf import PdfReader
        reader = PdfReader(BytesIO(file_bytes))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    if name.endswith(".docx"):
        from docx import Document as DocxDocument
        doc = DocxDocument(BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        lines = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append(" ".join(cells))
        return "\n".join(lines)
    # fall back to plain text
    return file_bytes.decode("utf-8", errors="ignore")


def chunk_text(text: str, chunk_size: int = 900, overlap: int = 150) -> List[str]:
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start = end - overlap
    return [c.strip() for c in chunks if len(c.strip()) > 50]


def build_tfidf_matrix(chunks: List[str]):
    """Used for retrieval if we ever need to fetch the most relevant chunks
    for a specific topic. Returns the vectorizer and matrix as plain lists
    so they are JSON/Mongo friendly."""
    vectorizer = TfidfVectorizer(stop_words="english", max_features=2000)
    matrix = vectorizer.fit_transform(chunks)
    return vectorizer, matrix


def most_relevant_chunks(query: str, chunks: List[str], vectorizer, matrix, top_k: int = 3) -> List[str]:
    query_vec = vectorizer.transform([query])
    scores = cosine_similarity(query_vec, matrix).flatten()
    top_idx = np.argsort(scores)[::-1][:top_k]
    return [chunks[i] for i in top_idx]


QUESTION_PROMPT = """You are writing a short quiz for a brand-new employee based on the document excerpt below.

Write {n} multiple choice questions in VERY SIMPLE, EASY English:
- Short sentences, everyday common words. Simplify any jargon or formal wording from the excerpt - do not copy complex phrasing.
- A question should be understandable in one quick read, like explaining it to someone with basic English.
- Ask about ONE clear fact per question. Never combine multiple facts, and never use negatives or trick wording (e.g. "which of these is NOT...").
- The correct answer must be directly and plainly stated in the excerpt - never something the reader has to guess or infer.
- Each question must have exactly 4 options, only one correct, and the 3 wrong options must be clearly and unambiguously wrong.

Return ONLY valid JSON, no other text, in this exact shape:
[
  {{"question": "...", "options": ["...", "...", "...", "..."], "correct_index": 0}}
]

Excerpt:
\"\"\"{excerpt}\"\"\"
"""


def generate_questions_for_chunk(chunk: str, n: int) -> List[dict]:
    if _client is None:
        raise RuntimeError(
            "OPENAI_API_KEY is not configured. Add it to .env to enable question generation."
        )
    resp = _client.chat.completions.create(
        model="gpt-4o-mini",
        max_tokens=1500,
        messages=[{"role": "user", "content": QUESTION_PROMPT.format(n=n, excerpt=chunk)}],
    )
    raw = resp.choices[0].message.content or ""
    raw = raw.strip().strip("`")
    if raw.lower().startswith("json"):
        raw = raw[4:]
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\[.*\]", raw, re.DOTALL)
        data = json.loads(match.group(0)) if match else []

    cleaned = []
    for item in data:
        if (
            isinstance(item, dict)
            and item.get("question")
            and isinstance(item.get("options"), list)
            and len(item["options"]) == 4
            and isinstance(item.get("correct_index"), int)
            and 0 <= item["correct_index"] <= 3
        ):
            # The model tends to always place the correct answer first, so
            # shuffle option order here rather than trust its positioning.
            options = [o.strip() for o in item["options"]]
            correct_answer = options[item["correct_index"]]
            random.shuffle(options)
            cleaned.append(
                {
                    "id": str(uuid.uuid4()),
                    "question": item["question"].strip(),
                    "options": options,
                    "correct_index": options.index(correct_answer),
                }
            )
    return cleaned


def build_question_pool(chunks: List[str], per_chunk: int = 3) -> List[dict]:
    """Calls the LLM once per chunk (small excerpts, not the whole document
    at once) so every part of the document gets read and gets a chance to
    contribute questions. No cap on the total - the pool size scales with
    how long the document is. A random 15 are picked from this pool for
    each candidate's actual quiz."""
    pool: List[dict] = []
    for chunk in chunks:
        try:
            pool.extend(generate_questions_for_chunk(chunk, per_chunk))
        except RuntimeError:
            raise
        except Exception:
            continue
    return pool
