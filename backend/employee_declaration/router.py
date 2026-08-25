# backend/employee_declaration/router.py
"""
Employee Declaration Form - every employee reads the fixed declaration
document and signs it once; two allowlisted employee codes (see db.py's
`admins` collection) get a read-only admin roster + Excel export.

Uses the platform's normal get_current_user - unlike HR Policy Quiz there's
no separate non-employee login to bridge from, so no extra JWT domain here.
"""
import base64
import html as html_lib
import re
from datetime import datetime, timezone
from io import BytesIO

import pytz
from fastapi import APIRouter, Depends, HTTPException, Response, status
from pydantic import BaseModel

from backend.auth import get_current_user
from backend.database import employee_details_collection
from backend.employee_declaration.db import submissions, admins
from backend.employee_declaration.html_render import DECLARATION_HTML

router = APIRouter(prefix="/employee-declaration", tags=["Employee Declaration"])

IST = pytz.timezone("Asia/Kolkata")


def now():
    return datetime.now(timezone.utc)


def _format_ist(dt) -> str:
    """submitted_at is stored as UTC (pymongo hands it back tz-naive) - the
    admin table/JS side convert to the browser's local time automatically,
    but the PDF/Excel exports are rendered server-side, so they need an
    explicit UTC -> IST conversion instead of just formatting the raw UTC
    value as if it were already local."""
    if not dt:
        return "-"
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST).strftime("%d %b %Y, %I:%M %p")


def _employee_name(empid: str) -> str:
    emp = employee_details_collection.find_one({"EmpID": empid})
    if not emp:
        return empid
    return emp.get("Emp Name") or emp.get("EmployeeName") or emp.get("Name") or empid


def _is_admin_empid(empid: str) -> bool:
    empid = empid.strip().upper()
    if admins.find_one({"empid": empid}):
        return True
    # Case-insensitive fallback, in case someone inserts a code manually
    # from Compass in lowercase or mixed case.
    return admins.find_one({"empid": {"$regex": f"^{empid}$", "$options": "i"}}) is not None


def _require_admin(current_user: str) -> str:
    empid = current_user.strip().upper()
    if not _is_admin_empid(empid):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for the Employee Declaration Admin dashboard")
    return empid


# ---------------------------------------------------------------------------
# Employee: view + sign + submit
# ---------------------------------------------------------------------------

@router.get("/document")
def get_document(current_user: str = Depends(get_current_user)):
    return {"html": DECLARATION_HTML}


@router.get("/status")
def get_status(current_user: str = Depends(get_current_user)):
    empid = current_user.strip().upper()
    record = submissions.find_one({"empid": empid})
    if not record:
        return {"submitted": False, "submitted_at": None, "employee_name": _employee_name(empid), "signature": None}
    return {
        "submitted": True,
        "submitted_at": record.get("submitted_at"),
        "employee_name": record.get("employee_name") or _employee_name(empid),
        "signature": record.get("signature"),
    }


class SubmitDeclarationRequest(BaseModel):
    signature: str


@router.post("/submit")
def submit_declaration(body: SubmitDeclarationRequest, current_user: str = Depends(get_current_user)):
    empid = current_user.strip().upper()

    if not body.signature or not body.signature.startswith("data:image"):
        raise HTTPException(status_code=422, detail="A signature is required before you can submit")

    if submissions.find_one({"empid": empid}):
        raise HTTPException(status_code=409, detail="You have already submitted this declaration")

    employee_name = _employee_name(empid)
    submitted_at = now()
    submissions.insert_one(
        {
            "empid": empid,
            "employee_name": employee_name,
            "signature": body.signature,
            "submitted_at": submitted_at,
        }
    )
    return {"success": True, "submitted_at": submitted_at, "employee_name": employee_name}


# ---------------------------------------------------------------------------
# Admin: full roster + export
# ---------------------------------------------------------------------------

@router.get("/admin/access-check")
def admin_access_check(current_user: str = Depends(get_current_user)):
    return {"is_admin": _is_admin_empid(current_user)}


def _display_name(emp: dict, empid: str) -> str:
    return emp.get("Emp Name") or emp.get("EmployeeName") or emp.get("Name") or empid


def _gather_roster() -> list[dict]:
    submitted_by_empid = {s["empid"]: s for s in submissions.find({})}
    rows = []
    for emp in employee_details_collection.find({}, {"EmpID": 1, "Emp Name": 1, "EmployeeName": 1, "Name": 1}):
        empid = (emp.get("EmpID") or "").strip().upper()
        if not empid:
            continue
        record = submitted_by_empid.get(empid)
        rows.append(
            {
                "empid": empid,
                "name": _display_name(emp, empid),
                "submitted": record is not None,
                "submitted_at": record.get("submitted_at") if record else None,
                "signature": record.get("signature") if record else None,
            }
        )
    rows.sort(key=lambda r: (not r["submitted"], r["empid"]))
    return rows


@router.get("/admin/submissions")
def admin_submissions(current_user: str = Depends(get_current_user)):
    _require_admin(current_user)
    return {"rows": _gather_roster()}


def _declaration_paragraphs() -> list[str]:
    """Strip mammoth's simple HTML down to plain paragraphs for the PDF."""
    text = re.sub(r"</(p|h1|h2|h3|h4|h5|h6|li|tr|div)>", "\n", DECLARATION_HTML, flags=re.I)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = html_lib.unescape(text)
    return [ln.strip() for ln in text.split("\n") if ln.strip()]


# fpdf2's core Helvetica font only supports Latin-1, but the source .docx
# (and Word docs in general) is full of "smart" punctuation outside that
# range - typing any of it straight into pdf.cell()/multi_cell() throws and
# was silently turning every PDF download into a 500.
_PDF_UNICODE_REPLACEMENTS = {
    "‘": "'", "’": "'", "‚": "'",
    "“": '"', "”": '"', "„": '"',
    "–": "-", "—": "-", "−": "-",
    "…": "...", " ": " ", "•": "-",
}


def _pdf_safe(text: str) -> str:
    for uni, ascii_ in _PDF_UNICODE_REPLACEMENTS.items():
        text = text.replace(uni, ascii_)
    return text.encode("latin-1", "replace").decode("latin-1")


def _generate_declaration_pdf(empid: str, name: str, submitted_at, signature: str | None) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    submitted_str = _format_ist(submitted_at)
    name = _pdf_safe(name)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Employee Declaration & Undertaking", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, "Code of Conduct, Confidentiality & Compliance", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, f"Employee: {name} ({empid})", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Submitted: {submitted_str}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    pdf.set_draw_color(180, 180, 180)
    pdf.line(pdf.get_x(), pdf.get_y(), pdf.get_x() + 190, pdf.get_y())
    pdf.ln(6)

    pdf.set_font("Helvetica", "", 9.5)
    for para in _declaration_paragraphs():
        try:
            pdf.multi_cell(0, 5.2, _pdf_safe(para), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(1.5)
        except Exception:
            continue

    pdf.ln(2)
    pdf.set_draw_color(180, 180, 180)
    pdf.line(pdf.get_x(), pdf.get_y(), pdf.get_x() + 190, pdf.get_y())
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Signed acknowledgement", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 9.5)
    pdf.multi_cell(
        0, 5.2,
        "I have read, understood, and agree to comply with the above Code of Conduct, "
        "Confidentiality & Compliance declaration.",
        new_x=XPos.LMARGIN, new_y=YPos.NEXT,
    )
    pdf.ln(4)

    col_w = 63
    pdf.set_font("Helvetica", "B", 8.5)
    pdf.set_fill_color(245, 245, 245)
    pdf.cell(col_w, 7, "Employee Name", border=1, fill=True)
    pdf.cell(col_w, 7, "Date", border=1, fill=True)
    pdf.cell(col_w, 7, "Signature", border=1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.set_font("Helvetica", "", 9)
    sig_row_h = 22
    pdf.cell(col_w, sig_row_h, name, border=1)
    pdf.cell(col_w, sig_row_h, submitted_str, border=1)
    sig_x, sig_y = pdf.get_x(), pdf.get_y()
    pdf.cell(col_w, sig_row_h, "", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    if signature and signature.startswith("data:image"):
        try:
            img_bytes = base64.b64decode(signature.split(",", 1)[1])
            pdf.image(BytesIO(img_bytes), x=sig_x + 2, y=sig_y + 2, w=col_w - 4, h=sig_row_h - 4, type="PNG")
        except Exception:
            pass

    return bytes(pdf.output())


@router.get("/admin/submissions/{empid}/pdf")
def admin_submission_pdf(empid: str, current_user: str = Depends(get_current_user)):
    _require_admin(current_user)
    empid = empid.strip().upper()

    record = submissions.find_one({"empid": empid})
    if not record:
        raise HTTPException(status_code=404, detail="No submission found for this employee")

    name = record.get("employee_name") or _employee_name(empid)
    try:
        pdf_bytes = _generate_declaration_pdf(empid, name, record.get("submitted_at"), record.get("signature"))
    except Exception as e:
        print(f"[DeclarationPDF] Failed to generate PDF for {empid}: {e}")
        raise HTTPException(status_code=500, detail="Could not generate the declaration PDF")

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=declaration-{empid}.pdf"},
    )


@router.get("/admin/export.xlsx")
def admin_export_xlsx(current_user: str = Depends(get_current_user)):
    _require_admin(current_user)

    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = _gather_roster()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Declarations"
    bold = Font(bold=True)

    sheet.append(["Employee ID", "Name", "Submitted", "Submitted At"])
    for cell in sheet[1]:
        cell.font = bold

    for r in rows:
        sheet.append([r["empid"], r["name"], "Yes" if r["submitted"] else "No", _format_ist(r["submitted_at"])])

    for col in sheet.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        sheet.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)

    buf = BytesIO()
    wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee-declaration-report.xlsx"},
    )
