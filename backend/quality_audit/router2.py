# backend/quality_audit/router.py  (v3 — improvements field, misses column, rich dashboard)
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from backend.auth import get_current_user
from backend.database import employee_details_collection
from backend.quality_audit.db import (
    qa_users_collection,
    qa_audit_collection,
    qa_temp_collection,
)

router = APIRouter(prefix="/quality-audit", tags=["Quality Audit"])

QA_ONLY_EMAILS = {"bhupen.shah@jhsassociates.in"}

SECTION_LABELS = {
    "section_a": "Audit Planning",
    "section_b": "Audit Execution",
    "section_c": "Audit Reporting",
    "section_d": "Quality Control",
    "section_e": "Optimization",
}


# ─── helpers ──────────────────────────────────────────────────────────────────

def _get_access_doc():
    return qa_users_collection.find_one({}) or {"user": [], "admin": []}


def _email_from_emp(emp_id: str) -> str:
    emp = employee_details_collection.find_one({"EmpID": emp_id.strip().upper()})
    if emp:
        return (emp.get("EMail") or emp.get("JHS Email") or "").lower().strip()
    return ""


def _check_qa_access(current_user: str):
    email  = _email_from_emp(current_user)
    doc    = _get_access_doc()
    users  = [e.lower().strip() for e in doc.get("user",  [])]
    admins = [e.lower().strip() for e in doc.get("admin", [])]
    is_admin   = email in admins
    is_user    = email in users
    is_qa_only = email in QA_ONLY_EMAILS
    if not (is_admin or is_user):
        raise HTTPException(403, "Access denied: not in Quality Audit user list")
    return email, is_admin, is_qa_only


# ─── Pydantic ─────────────────────────────────────────────────────────────────

class AuditPayload(BaseModel):
    project_id:            Optional[str] = ""
    project_name:          Optional[str] = ""
    client_name:           Optional[str] = ""
    type_of_audit:         Optional[str] = ""
    audit_period_from:     Optional[str] = ""
    audit_period_to:       Optional[str] = ""
    audit_given_by_name:   str
    audit_given_by_emp_id: Optional[str] = ""
    project_tl:            str
    project_pnd:           str
    audit_date:            Optional[str] = ""
    section_a: List[Dict[str, Any]] = []
    section_b: List[Dict[str, Any]] = []
    section_c: List[Dict[str, Any]] = []
    section_d: List[Dict[str, Any]] = []
    section_e: List[Dict[str, Any]] = []
    improvements: Optional[str] = ""       # ← NEW
    total_score:  Optional[float] = 0
    ideal_score:  Optional[float] = 0
    scaled_total: Optional[float] = 0


# ─── access check ─────────────────────────────────────────────────────────────

@router.get("/access-check")
async def access_check(current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    return {"has_access": True, "is_admin": is_admin, "is_qa_only": is_qa_only, "email": email}


# ─── dropdowns ────────────────────────────────────────────────────────────────

@router.get("/dropdowns")
async def get_dropdowns(current_user: str = Depends(get_current_user)):
    _check_qa_access(current_user)
    mgr_set: dict[str, str] = {}
    for doc in employee_details_collection.find(
        {"ReportingEmpName": {"$exists": True, "$ne": ""}},
        {"ReportingEmpName": 1, "ReportingEmpCode": 1, "_id": 0}
    ):
        n = (doc.get("ReportingEmpName") or "").strip()
        c = (doc.get("ReportingEmpCode") or "").strip()
        if n: mgr_set[n] = c

    managers = sorted([{"name": n, "code": c} for n, c in mgr_set.items()], key=lambda x: x["name"])

    pnd_set = set()
    for doc in employee_details_collection.find({"PnD": {"$exists": True, "$ne": ""}}, {"PnD": 1, "_id": 0}):
        v = (doc.get("PnD") or "").strip()
        if v: pnd_set.add(v)
    if not pnd_set:
        for doc in employee_details_collection.find({"PartnerName": {"$exists": True, "$ne": ""}}, {"PartnerName": 1, "_id": 0}):
            v = (doc.get("PartnerName") or "").strip()
            if v: pnd_set.add(v)

    return {"managers": managers, "pnds": sorted(list(pnd_set))}


# ─── save draft ───────────────────────────────────────────────────────────────

@router.post("/save-draft")
async def save_draft(payload: AuditPayload, current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    if is_qa_only:
        raise HTTPException(403, "Dashboard-only access")
    doc = payload.dict()
    doc.update({"saved_by": current_user, "saved_email": email, "saved_at": datetime.utcnow(), "status": "draft"})
    qa_temp_collection.update_one({"saved_by": current_user}, {"$set": doc}, upsert=True)
    return {"success": True, "message": "Draft saved"}


@router.get("/load-draft")
async def load_draft(current_user: str = Depends(get_current_user)):
    _check_qa_access(current_user)
    doc = qa_temp_collection.find_one({"saved_by": current_user})
    if not doc: return {"draft": None}
    doc["_id"] = str(doc["_id"])
    return {"draft": doc}


# ─── submit ───────────────────────────────────────────────────────────────────

@router.post("/submit")
async def submit_audit(payload: AuditPayload, current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    if is_qa_only:
        raise HTTPException(403, "Dashboard-only access")
    doc = payload.dict()
    doc.update({"submitted_by": current_user, "submitted_email": email,
                "submitted_at": datetime.utcnow(), "status": "submitted"})
    result = qa_audit_collection.insert_one(doc)
    qa_temp_collection.delete_one({"saved_by": current_user})
    return {"success": True, "message": "Audit submitted", "audit_id": str(result.inserted_id)}


# ─── history ──────────────────────────────────────────────────────────────────

@router.get("/history")
async def get_history(current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    show_all = is_admin or is_qa_only
    query    = {} if show_all else {"submitted_by": current_user}
    cursor   = qa_audit_collection.find(query, {
        "_id": 1, "audit_date": 1, "project_tl": 1, "client_name": 1,
        "project_name": 1, "submitted_at": 1, "total_score": 1,
        "ideal_score": 1, "scaled_total": 1, "submitted_by": 1,
    }).sort("submitted_at", -1)

    items = []
    for doc in cursor:
        sat = doc.get("submitted_at")
        items.append({
            "id":           str(doc["_id"]),
            "audit_date":   doc.get("audit_date", ""),
            "project_tl":   doc.get("project_tl", ""),
            "client_name":  doc.get("client_name", ""),
            "project_name": doc.get("project_name", ""),
            "submitted_at": sat.isoformat() if isinstance(sat, datetime) else str(sat or ""),
            "total_score":  doc.get("total_score", 0),
            "ideal_score":  doc.get("ideal_score", 0),
            "scaled_total": doc.get("scaled_total", 0),
            "submitted_by": doc.get("submitted_by", ""),
        })
    return {"history": items, "is_admin": is_admin or is_qa_only}


# ─── single audit detail ──────────────────────────────────────────────────────

@router.get("/audit/{audit_id}")
async def get_audit(audit_id: str, current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    try: oid = ObjectId(audit_id)
    except: raise HTTPException(400, "Invalid audit ID")
    doc = qa_audit_collection.find_one({"_id": oid})
    if not doc: raise HTTPException(404, "Not found")
    if not (is_admin or is_qa_only) and doc.get("submitted_by") != current_user:
        raise HTTPException(403, "Access denied")
    doc["_id"] = str(doc["_id"])
    if isinstance(doc.get("submitted_at"), datetime):
        doc["submitted_at"] = doc["submitted_at"].isoformat()
    return {"audit": doc}


# ─── dashboard stats (rich) ───────────────────────────────────────────────────

def _safe_float(v, default=0.0):
    """Convert any value to float safely."""
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def _parse_sat(sat):
    """Return a datetime regardless of whether sat is already datetime or string."""
    if isinstance(sat, datetime):
        return sat
    if isinstance(sat, str) and sat:
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(sat[:26], fmt)
            except ValueError:
                continue
    return None


@router.get("/dashboard-stats")
async def dashboard_stats(current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    if not (is_admin or is_qa_only):
        raise HTTPException(403, "Dashboard restricted to admins")

    all_docs = list(qa_audit_collection.find({}))

    records            = []
    tl_scores          = {}
    pnd_scores         = {}
    client_scores      = {}
    month_data         = {}   # "YYYY-MM" -> {label, count}
    section_scores     = {k: {"score": 0.0, "ideal": 0.0, "count": 0} for k in SECTION_LABELS}
    particular_scores  = {}
    misses_list        = []

    for doc in all_docs:
        sat = _parse_sat(doc.get("submitted_at"))

        tl  = (doc.get("project_tl")  or "").strip()
        pnd = (doc.get("project_pnd") or "").strip()
        cli = (doc.get("client_name") or doc.get("project_name") or "").strip()
        sc  = _safe_float(doc.get("total_score"))
        isc = _safe_float(doc.get("ideal_score"))
        # Recompute scaled from raw scores — handles legacy & imported records
        sca = round(sc * 100 / isc, 2) if isc > 0 else _safe_float(doc.get("scaled_total"))

        # Month — sortable key "YYYY-MM", human label "Mon YYYY"
        if sat:
            mon_key   = sat.strftime("%Y-%m")
            mon_label = sat.strftime("%b %Y")
        else:
            mon_key   = "0000-00"
            mon_label = "Unknown"
        if mon_key not in month_data:
            month_data[mon_key] = {"label": mon_label, "count": 0}
        month_data[mon_key]["count"] += 1

        if tl:  tl_scores.setdefault(tl, []).append(sca)
        if pnd: pnd_scores.setdefault(pnd, []).append(sca)
        if cli: client_scores.setdefault(cli, []).append(sca)

        for sec_key, sec_label in SECTION_LABELS.items():
            items = doc.get(sec_key) or []
            if not isinstance(items, list):
                continue
            for it in items:
                s   = _safe_float(it.get("score"))
                id_ = _safe_float(it.get("ideal_score"))
                section_scores[sec_key]["score"] += s
                section_scores[sec_key]["ideal"] += id_
                section_scores[sec_key]["count"] += 1

                p = (it.get("particular") or "").strip()
                if p:
                    particular_scores.setdefault(p, {"score": 0.0, "ideal": 0.0, "count": 0})
                    particular_scores[p]["score"] += s
                    particular_scores[p]["ideal"] += id_
                    particular_scores[p]["count"] += 1

                ms = (it.get("misses") or "").strip()
                if ms:
                    misses_list.append({
                        "particular": p,
                        "section":    sec_label,
                        "misses":     ms,
                        "tl":         tl,
                        "client":     cli,
                        "date":       doc.get("audit_date") or (sat.strftime("%d %b %Y") if sat else ""),
                    })

        records.append({
            "id":           str(doc["_id"]),
            "audit_date":   doc.get("audit_date", ""),
            "project_tl":   tl,
            "project_pnd":  pnd,
            "client_name":  cli,
            "submitted_at": sat.isoformat() if sat else "",
            "total_score":  sc,
            "ideal_score":  isc,
            "scaled_total": sca,
            "submitted_by": doc.get("submitted_by", ""),
            "improvements": doc.get("improvements", ""),
        })

    def avg(lst): return round(sum(lst) / len(lst), 1) if lst else 0

    tl_board = sorted(
        [{"name": k, "avg": avg(v), "count": len(v)} for k, v in tl_scores.items()],
        key=lambda x: x["avg"], reverse=True,
    )
    pnd_board = sorted(
        [{"name": k, "avg": avg(v), "count": len(v)} for k, v in pnd_scores.items()],
        key=lambda x: x["avg"], reverse=True,
    )
    client_board = sorted(
        [{"name": k, "avg": avg(v), "count": len(v)} for k, v in client_scores.items()],
        key=lambda x: x["avg"], reverse=True,
    )

    # Month trend — sorted chronologically
    month_trend = [
        {"month": v["label"], "count": v["count"]}
        for k, v in sorted(month_data.items())
        if k != "0000-00"
    ]

    section_breakdown = []
    for sec_key, sec_label in SECTION_LABELS.items():
        sd  = section_scores[sec_key]
        pct = round(sd["score"] / sd["ideal"] * 100, 1) if sd["ideal"] > 0 else 0
        section_breakdown.append({
            "section": sec_label,
            "score":   round(sd["score"], 1),
            "ideal":   round(sd["ideal"], 1),
            "pct":     pct,
        })

    particular_breakdown = []
    for p, pd in particular_scores.items():
        pct = round(pd["score"] / pd["ideal"] * 100, 1) if pd["ideal"] > 0 else 0
        particular_breakdown.append({
            "particular": p[:80], "score": round(pd["score"], 1),
            "ideal": round(pd["ideal"], 1), "pct": pct, "count": pd["count"],
        })
    particular_breakdown.sort(key=lambda x: x["pct"])

    top5_clients    = client_board[:5]
    bottom5_clients = sorted(client_board, key=lambda x: x["avg"])[:5]
    overall_avg     = avg([r["scaled_total"] for r in records])
    overall_ideal   = round(sum(r["ideal_score"] for r in records) / len(records), 1) if records else 0

    return {
        "total_audits":         len(records),
        "unique_employees":     len(tl_scores),
        "unique_clients":       len(client_scores),
        "overall_avg_score":    overall_avg,
        "overall_ideal_score":  overall_ideal,
        "records":              records,
        "tl_board":             tl_board,
        "pnd_board":            pnd_board,
        "client_board":         client_board,
        "top5_clients":         top5_clients,
        "bottom5_clients":      bottom5_clients,
        "month_trend":          month_trend,
        "section_breakdown":    section_breakdown,
        "particular_breakdown": particular_breakdown,
        "misses_list":          misses_list,
    }


# ─── export ───────────────────────────────────────────────────────────────────

@router.get("/export/{audit_id}")
async def export_audit(audit_id: str, current_user: str = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    email, is_admin, is_qa_only = _check_qa_access(current_user)
    try: oid = ObjectId(audit_id)
    except: raise HTTPException(400, "Invalid ID")
    doc = qa_audit_collection.find_one({"_id": oid})
    if not doc: raise HTTPException(404, "Not found")
    if not (is_admin or is_qa_only) and doc.get("submitted_by") != current_user:
        raise HTTPException(403, "Access denied")

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"

    PURPLE = "7B2D8B"; LIGHT_PURPLE = "E8D5F5"; BLUE_HDR = "1F4E79"
    ORANGE = "C65911"; YELLOW = "FFF2CC"; LIGHT_GREEN = "E2EFDA"
    LIGHT_GREY = "F2F2F2"; WHITE = "FFFFFF"; RED_LIGHT = "FFCCCC"

    def cs(ws, row, col, value="", bold=False, bg=None, fc="000000",
           size=10, wrap=False, ha="left", va="center"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, size=size, color=fc)
        if bg: c.fill = PatternFill("solid", start_color=bg, end_color=bg)
        c.alignment = Alignment(wrap_text=wrap, horizontal=ha, vertical=va)
        thin = Side(style="thin", color="BBBBBB")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    def sec_hdr(ws, row, label, bg=BLUE_HDR, ncols=7):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", start_color=bg, end_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 28  # Misses/Serious
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 9   # (spare)

    # Title
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    tc = ws.cell(row=r, column=1, value="JHS QUALITY AUDIT CHECKLIST")
    tc.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    tc.fill = PatternFill("solid", start_color=PURPLE, end_color=PURPLE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 28

    info_rows = [
        ("Project ID",            doc.get("project_id", "")),
        ("Project Name",          doc.get("project_name", "")),
        ("Client Name",           doc.get("client_name", "")),
        ("Type of Audit",         doc.get("type_of_audit", "")),
        ("Audit Period",          f"From: {doc.get('audit_period_from','')}   To: {doc.get('audit_period_to','')}"),
        ("Quality Audit Given By",f"{doc.get('audit_given_by_name','')}  (Emp ID: {doc.get('audit_given_by_emp_id','')})"),
        ("Project TL / Manager",  doc.get("project_tl", "")),
        ("Project PnD",           doc.get("project_pnd", "")),
        ("Audit Date",            doc.get("audit_date", "")),
    ]
    for label, val in info_rows:
        r += 1
        ws.row_dimensions[r].height = 15
        cs(ws, r, 1, label, bold=True, bg=LIGHT_PURPLE, fc=PURPLE)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        cs(ws, r, 2, val, bg=WHITE)

    # Column headers
    r += 1
    ws.row_dimensions[r].height = 20
    for ci, h in enumerate(["Particulars","YES / NO / NA","Evidence / Remark","Misses / Serious","Score","Ideal Score"], 1):
        cs(ws, r, ci, h, bold=True, bg=BLUE_HDR, fc=WHITE, ha="center")

    sections_cfg = [
        ("section_a", "(A) Audit Planning",  BLUE_HDR),
        ("section_b", "(B) Audit Execution", "375623"),
        ("section_c", "(C) Audit Reporting", ORANGE),
        ("section_d", "(D) Quality Control", PURPLE),
        ("section_e", "(E) Optimization",    "1E6B3C"),
    ]

    r += 1
    for sec_key, sec_label, sec_color in sections_cfg:
        sec_hdr(ws, r, sec_label, bg=sec_color, ncols=6)
        r += 1
        for item in doc.get(sec_key, []):
            ws.row_dimensions[r].height = 28
            resp = item.get("response", "")
            rb   = "E2EFDA" if resp == "YES" else "FFCCCC" if resp == "NO" else LIGHT_GREY
            cs(ws, r, 1, item.get("particular",""), bg=LIGHT_GREY, wrap=True, va="top")
            cs(ws, r, 2, resp, bg=rb, ha="center", bold=True)
            cs(ws, r, 3, item.get("evidence",""), bg=WHITE, wrap=True, va="top")
            cs(ws, r, 4, item.get("misses",""), bg="FFF8E1", wrap=True, va="top")
            cs(ws, r, 5, item.get("score",""), bg=YELLOW, ha="center", bold=True)
            cs(ws, r, 6, item.get("ideal_score",""), bg=LIGHT_GREEN, ha="center")
            r += 1

    # Improvements
    improvements = (doc.get("improvements") or "").strip()
    if improvements:
        r += 1
        sec_hdr(ws, r, "Improvements (if any)", bg="1565C0", ncols=6)
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=6)
        c = ws.cell(row=r, column=1, value=improvements)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.fill = PatternFill("solid", start_color="E3F2FD", end_color="E3F2FD")
        ws.row_dimensions[r].height = 60
        r += 4

    # Totals
    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cs(ws, r, 1, "TOTAL", bold=True, bg=BLUE_HDR, fc=WHITE, ha="right")
    cs(ws, r, 5, doc.get("total_score", 0), bold=True, bg=YELLOW, ha="center")
    cs(ws, r, 6, doc.get("ideal_score", 0), bold=True, bg=LIGHT_GREEN, ha="center")
    r += 1
    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cs(ws, r, 1, "SCALED TOTAL (out of 100)", bold=True, bg=PURPLE, fc=WHITE, ha="right")
    cs(ws, r, 5, round(doc.get("scaled_total", 0), 2), bold=True, bg=LIGHT_PURPLE, ha="center", fc=PURPLE)
    cs(ws, r, 6, 100, bold=True, bg=LIGHT_GREEN, ha="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe  = (doc.get("client_name") or "Audit").replace(" ","_")[:30]
    dstr  = (doc.get("audit_date")  or "")[:10].replace("/","-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="QA_{safe}_{dstr}.xlsx"'},
    )
