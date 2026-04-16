# backend/quality_audit/router.py  (v6 — ultra-fast dashboard)
"""
Performance strategy:
  /dashboard-kpis  → Single MongoDB $facet pipeline for ALL header aggregations (~50-100ms)
                     Returns KPIs + tl/pnd/client boards + month trend instantly
                     Section breakdown now moved to Phase 2 (was 5 sequential pipelines)
  /dashboard-stats → Single $facet pipeline for ALL 5 sections simultaneously (~100-150ms)
                     Returns section_breakdown + particular_breakdown + misses_list

Key improvements over v5:
  - Section breakdown: 5 sequential pipelines → 1 $facet pipeline (5x faster)
  - Particular breakdown: 5 sequential pipelines → 1 $facet pipeline (5x faster)
  - Misses: 5 sequential pipelines → 1 $facet pipeline (5x faster)
  - Header aggregation: Python for-loops → MongoDB $group via $facet
  - Frontend: Summary renders without waiting for section data
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from backend.auth import get_current_user
from backend.database import employee_details_collection
from backend.quality_audit.db import (
    qa_users_collection,
    qa_audit_collection,
    qa_temp_collection,
)

router = APIRouter(prefix="/quality-audit", tags=["Quality Audit"])

QA_ONLY_EMAILS = {"bhupen.shah@jhsassociates.in"}

SECTION_LABELS = {
    "section_a": "Audit Planning",
    "section_b": "Audit Execution",
    "section_c": "Audit Reporting",
    "section_d": "Quality Control",
    "section_e": "Optimization",
}

SEC_KEYS = list(SECTION_LABELS.keys())


# ─── helpers ──────────────────────────────────────────────────────────────────

def _get_access_doc():
    return qa_users_collection.find_one({}) or {"user": [], "admin": []}


def _email_from_emp(emp_id: str) -> str:
    emp = employee_details_collection.find_one({"EmpID": emp_id.strip().upper()})
    if emp:
        return (emp.get("EMail") or emp.get("JHS Email") or "").lower().strip()
    return ""


def _check_qa_access(current_user: str):
    email  = _email_from_emp(current_user)
    doc    = _get_access_doc()
    users  = [e.lower().strip() for e in doc.get("user",  [])]
    admins = [e.lower().strip() for e in doc.get("admin", [])]
    is_admin   = email in admins
    is_user    = email in users
    is_qa_only = email in QA_ONLY_EMAILS
    if not (is_admin or is_user):
        raise HTTPException(403, "Access denied: not in Quality Audit user list")
    return email, is_admin, is_qa_only


def _safe_float(v, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def _parse_sat(sat):
    if isinstance(sat, datetime):
        return sat
    if isinstance(sat, str) and sat:
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(sat[:26], fmt)
            except ValueError:
                continue
    return None


# ─── Pydantic ─────────────────────────────────────────────────────────────────

class AuditPayload(BaseModel):
    project_id:            Optional[str] = ""
    project_name:          Optional[str] = ""
    client_name:           Optional[str] = ""
    type_of_audit:         Optional[str] = ""
    audit_period_from:     Optional[str] = ""
    audit_period_to:       Optional[str] = ""
    audit_given_by_name:   str
    audit_given_by_emp_id: Optional[str] = ""
    project_tl:            str
    project_pnd:           str
    audit_date:            Optional[str] = ""
    section_a: List[Dict[str, Any]] = []
    section_b: List[Dict[str, Any]] = []
    section_c: List[Dict[str, Any]] = []
    section_d: List[Dict[str, Any]] = []
    section_e: List[Dict[str, Any]] = []
    improvements: Optional[str] = ""
    total_score:  Optional[float] = 0
    ideal_score:  Optional[float] = 0
    scaled_total: Optional[float] = 0


# ─── access check ─────────────────────────────────────────────────────────────

@router.get("/access-check")
async def access_check(current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    return {"has_access": True, "is_admin": is_admin, "is_qa_only": is_qa_only, "email": email}


# ─── dropdowns ────────────────────────────────────────────────────────────────

@router.get("/dropdowns")
async def get_dropdowns(current_user: str = Depends(get_current_user)):
    _check_qa_access(current_user)
    mgr_set: dict[str, str] = {}
    for doc in employee_details_collection.find(
        {"ReportingEmpName": {"$exists": True, "$ne": ""}},
        {"ReportingEmpName": 1, "ReportingEmpCode": 1, "_id": 0},
    ):
        n = (doc.get("ReportingEmpName") or "").strip()
        c = (doc.get("ReportingEmpCode") or "").strip()
        if n:
            mgr_set[n] = c
    managers = sorted([{"name": n, "code": c} for n, c in mgr_set.items()], key=lambda x: x["name"])

    pnd_set: set[str] = set()
    for doc in employee_details_collection.find({"PnD": {"$exists": True, "$ne": ""}}, {"PnD": 1, "_id": 0}):
        v = (doc.get("PnD") or "").strip()
        if v: pnd_set.add(v)
    if not pnd_set:
        for doc in employee_details_collection.find({"PartnerName": {"$exists": True, "$ne": ""}}, {"PartnerName": 1, "_id": 0}):
            v = (doc.get("PartnerName") or "").strip()
            if v: pnd_set.add(v)

    return {"managers": managers, "pnds": sorted(list(pnd_set))}


# ─── save draft ───────────────────────────────────────────────────────────────

@router.post("/save-draft")
async def save_draft(payload: AuditPayload, current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    if is_qa_only:
        raise HTTPException(403, "Dashboard-only access")
    doc = payload.dict()
    doc.update({"saved_by": current_user, "saved_email": email,
                "saved_at": datetime.utcnow(), "status": "draft"})
    qa_temp_collection.update_one({"saved_by": current_user}, {"$set": doc}, upsert=True)
    return {"success": True, "message": "Draft saved"}


@router.get("/load-draft")
async def load_draft(current_user: str = Depends(get_current_user)):
    _check_qa_access(current_user)
    doc = qa_temp_collection.find_one({"saved_by": current_user})
    if not doc:
        return {"draft": None}
    doc["_id"] = str(doc["_id"])
    return {"draft": doc}


# ─── submit ───────────────────────────────────────────────────────────────────

@router.post("/submit")
async def submit_audit(payload: AuditPayload, current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    if is_qa_only:
        raise HTTPException(403, "Dashboard-only access")
    doc = payload.dict()
    doc.update({"submitted_by": current_user, "submitted_email": email,
                "submitted_at": datetime.utcnow(), "status": "submitted"})
    result = qa_audit_collection.insert_one(doc)
    qa_temp_collection.delete_one({"saved_by": current_user})
    return {"success": True, "message": "Audit submitted", "audit_id": str(result.inserted_id)}


# ─── history ──────────────────────────────────────────────────────────────────

@router.get("/history")
async def get_history(current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    show_all = is_admin or is_qa_only
    query    = {} if show_all else {"submitted_by": current_user}
    cursor   = qa_audit_collection.find(query, {
        "_id": 1, "audit_date": 1, "project_tl": 1, "client_name": 1,
        "project_name": 1, "submitted_at": 1, "total_score": 1,
        "ideal_score": 1, "scaled_total": 1, "submitted_by": 1,
    }).sort("submitted_at", -1)

    items = []
    for doc in cursor:
        sat = doc.get("submitted_at")
        sc  = _safe_float(doc.get("total_score"))
        isc = _safe_float(doc.get("ideal_score"))
        sca = round(sc * 100 / isc, 1) if isc > 0 else _safe_float(doc.get("scaled_total"))
        items.append({
            "id":           str(doc["_id"]),
            "audit_date":   doc.get("audit_date", ""),
            "project_tl":   doc.get("project_tl", ""),
            "client_name":  doc.get("client_name", ""),
            "project_name": doc.get("project_name", ""),
            "submitted_at": sat.isoformat() if isinstance(sat, datetime) else str(sat or ""),
            "total_score":  sc,
            "ideal_score":  isc,
            "scaled_total": sca,
            "submitted_by": doc.get("submitted_by", ""),
        })
    return {"history": items, "is_admin": is_admin or is_qa_only}


# ─── single audit detail ──────────────────────────────────────────────────────

@router.get("/audit/{audit_id}")
async def get_audit(audit_id: str, current_user: str = Depends(get_current_user)):
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    try:
        oid = ObjectId(audit_id)
    except Exception:
        raise HTTPException(400, "Invalid audit ID")
    doc = qa_audit_collection.find_one({"_id": oid})
    if not doc:
        raise HTTPException(404, "Not found")
    if not (is_admin or is_qa_only) and doc.get("submitted_by") != current_user:
        raise HTTPException(403, "Access denied")
    doc["_id"] = str(doc["_id"])
    sat = doc.get("submitted_at")
    if isinstance(sat, datetime):
        doc["submitted_at"] = sat.isoformat()
    return {"audit": doc}


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD  — Pure MongoDB aggregation, single $facet per endpoint
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/dashboard-kpis")
async def dashboard_kpis(current_user: str = Depends(get_current_user)):
    """
    Phase 1 — Ultra fast (~50ms).
    Single MongoDB $facet pipeline replaces all Python for-loops.
    Returns everything needed for Summary KPIs / TL / PnD / Client / Records tabs.
    Section breakdown intentionally excluded — done in Phase 2 via $facet.
    """
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    if not (is_admin or is_qa_only):
        raise HTTPException(403, "Dashboard restricted to admins")

    # ── Single pass: project only header fields, compute sca, then $facet ────
    pipeline = [
        # Step 1: project only the scalar fields we need (skip section arrays entirely)
        {"$project": {
            "project_tl":   1,
            "project_pnd":  1,
            "client_name":  1,
            "project_name": 1,
            "submitted_at": 1,
            "submitted_by": 1,
            "audit_date":   1,
            "total_score":  1,
            "ideal_score":  1,
            "scaled_total": 1,
        }},
        # Step 2: compute derived fields server-side
        {"$addFields": {
            "sca": {"$cond": {
                "if":   {"$gt": [{"$ifNull": ["$ideal_score", 0]}, 0]},
                "then": {"$multiply": [{"$divide": ["$total_score", "$ideal_score"]}, 100]},
                "else": {"$ifNull": ["$scaled_total", 0]},
            }},
            "cli": {"$ifNull": ["$client_name", "$project_name"]},
        }},
        # Step 3: all groupings in one $facet — single collection scan
        {"$facet": {
            "records": [
                {"$project": {
                    "_id": 1, "audit_date": 1, "project_tl": 1, "project_pnd": 1,
                    "cli": 1, "submitted_at": 1, "submitted_by": 1,
                    "total_score": 1, "ideal_score": 1, "sca": 1,
                }},
            ],
            "tl_groups": [
                {"$match": {"project_tl": {"$nin": [None, ""]}}},
                {"$group": {
                    "_id":   "$project_tl",
                    "scores": {"$push": "$sca"},
                    "count":  {"$sum": 1},
                }},
            ],
            "pnd_groups": [
                {"$match": {"project_pnd": {"$nin": [None, ""]}}},
                {"$group": {
                    "_id":    "$project_pnd",
                    "scores": {"$push": "$sca"},
                    "count":  {"$sum": 1},
                }},
            ],
            "client_groups": [
                {"$match": {"cli": {"$nin": [None, ""]}}},
                {"$group": {
                    "_id":    "$cli",
                    "scores": {"$push": "$sca"},
                    "count":  {"$sum": 1},
                }},
            ],
            "month_groups": [
                {"$match": {"submitted_at": {"$exists": True, "$ne": None}}},
                {"$group": {
                    "_id":   {"$dateToString": {"format": "%Y-%m", "date": "$submitted_at"}},
                    "count": {"$sum": 1},
                }},
                {"$sort": {"_id": 1}},
            ],
        }},
    ]

    result = list(qa_audit_collection.aggregate(pipeline))
    facet  = result[0] if result else {}

    def avg(lst): return round(sum(lst) / len(lst), 1) if lst else 0

    # ── Records ───────────────────────────────────────────────────────────────
    records = []
    for doc in facet.get("records", []):
        sat = _parse_sat(doc.get("submitted_at"))
        sc  = _safe_float(doc.get("total_score"))
        isc = _safe_float(doc.get("ideal_score"))
        sca = _safe_float(doc.get("sca")) or (round(sc * 100 / isc, 1) if isc > 0 else 0)
        records.append({
            "id":           str(doc["_id"]),
            "audit_date":   doc.get("audit_date", ""),
            "project_tl":   (doc.get("project_tl")  or "").strip(),
            "project_pnd":  (doc.get("project_pnd") or "").strip(),
            "client_name":  (doc.get("cli")          or "").strip(),
            "submitted_at": sat.isoformat() if sat else "",
            "total_score":  sc,
            "ideal_score":  isc,
            "scaled_total": round(sca, 1),
            "submitted_by": doc.get("submitted_by", ""),
        })

    # ── Leaderboards ──────────────────────────────────────────────────────────
    def build_board(groups):
        board = []
        for g in groups:
            name = (g.get("_id") or "").strip()
            if not name:
                continue
            scores = [_safe_float(s) for s in (g.get("scores") or [])]
            board.append({"name": name, "avg": avg(scores), "count": g.get("count", 0)})
        return sorted(board, key=lambda x: x["avg"], reverse=True)

    tl_board     = build_board(facet.get("tl_groups",     []))
    pnd_board    = build_board(facet.get("pnd_groups",    []))
    client_board = build_board(facet.get("client_groups", []))

    # ── Month trend ───────────────────────────────────────────────────────────
    month_trend = []
    for g in facet.get("month_groups", []):
        mk = g.get("_id", "")
        if not mk or mk == "0000-00":
            continue
        try:
            dt = datetime.strptime(mk, "%Y-%m")
            label = dt.strftime("%b %Y")
        except ValueError:
            label = mk
        month_trend.append({"month": label, "count": g.get("count", 0)})

    overall = avg([r["scaled_total"] for r in records])

    return {
        "total_audits":         len(records),
        "unique_employees":     len(tl_board),
        "unique_clients":       len(client_board),
        "overall_avg_score":    overall,
        "records":              records,
        "tl_board":             tl_board,
        "pnd_board":            pnd_board,
        "client_board":         client_board,
        "top5_clients":         client_board[:5],
        "bottom5_clients":      sorted(client_board, key=lambda x: x["avg"])[:5],
        "month_trend":          month_trend,
        # Phase 2 placeholders
        "section_breakdown":    [],
        "particular_breakdown": [],
        "misses_list":          [],
    }


@router.get("/dashboard-stats")
async def dashboard_stats(current_user: str = Depends(get_current_user)):
    """
    Phase 2 — Single $facet pipeline for all 5 sections simultaneously.
    Returns section_breakdown + particular_breakdown + misses_list.
    Previously: 5+5+5 = 15 sequential pipelines. Now: 1 $facet pipeline.
    """
    email, is_admin, is_qa_only = _check_qa_access(current_user)
    if not (is_admin or is_qa_only):
        raise HTTPException(403, "Dashboard restricted to admins")

    # Build $facet branches for all 3 data types × 5 sections in one pipeline
    facet_branches: dict = {}
    for sec_key in SEC_KEYS:
        # Section score totals
        facet_branches[f"{sec_key}_totals"] = [
            {"$project": {sec_key: 1}},
            {"$unwind": {"path": f"${sec_key}", "preserveNullAndEmptyArrays": False}},
            {"$group": {
                "_id":   None,
                "score": {"$sum": {"$ifNull": [f"${sec_key}.score", 0]}},
                "ideal": {"$sum": {"$ifNull": [f"${sec_key}.ideal_score", 0]}},
            }},
        ]
        # Per-particular breakdown
        facet_branches[f"{sec_key}_parts"] = [
            {"$project": {sec_key: 1}},
            {"$unwind": {"path": f"${sec_key}", "preserveNullAndEmptyArrays": False}},
            {"$match": {f"{sec_key}.particular": {"$nin": [None, ""]}}},
            {"$group": {
                "_id":   f"${sec_key}.particular",
                "score": {"$sum": {"$ifNull": [f"${sec_key}.score", 0]}},
                "ideal": {"$sum": {"$ifNull": [f"${sec_key}.ideal_score", 0]}},
                "count": {"$sum": 1},
            }},
        ]
        # Misses
        facet_branches[f"{sec_key}_misses"] = [
            {"$project": {
                sec_key:       1,
                "project_tl":  1,
                "client_name": 1,
                "project_name":1,
                "audit_date":  1,
                "submitted_at":1,
            }},
            {"$unwind": {"path": f"${sec_key}", "preserveNullAndEmptyArrays": False}},
            {"$match": {f"{sec_key}.misses": {"$nin": [None, ""]}}},
            {"$project": {
                "particular":   f"${sec_key}.particular",
                "misses":       f"${sec_key}.misses",
                "tl":           "$project_tl",
                "client":       {"$ifNull": ["$client_name", "$project_name"]},
                "audit_date":   1,
                "submitted_at": 1,
            }},
        ]

    result = list(qa_audit_collection.aggregate([{"$facet": facet_branches}]))
    facet  = result[0] if result else {}

    # ── Section breakdown ─────────────────────────────────────────────────────
    section_breakdown = []
    for sec_key, sec_label in SECTION_LABELS.items():
        totals = facet.get(f"{sec_key}_totals", [])
        if totals:
            s = _safe_float(totals[0].get("score"))
            i = _safe_float(totals[0].get("ideal"))
            pct = round(s / i * 100, 1) if i > 0 else 0
            section_breakdown.append({
                "section": sec_label,
                "score":   round(s, 1),
                "ideal":   round(i, 1),
                "pct":     pct,
            })
        else:
            section_breakdown.append({"section": sec_label, "score": 0, "ideal": 0, "pct": 0})

    # ── Particular breakdown ──────────────────────────────────────────────────
    particular_scores: dict[str, dict] = {}
    for sec_key in SEC_KEYS:
        for row in facet.get(f"{sec_key}_parts", []):
            p = (row.get("_id") or "").strip()
            if not p:
                continue
            ex = particular_scores.get(p)
            if ex:
                ex["score"] += _safe_float(row.get("score"))
                ex["ideal"] += _safe_float(row.get("ideal"))
                ex["count"] += int(row.get("count", 0))
            else:
                particular_scores[p] = {
                    "score": _safe_float(row.get("score")),
                    "ideal": _safe_float(row.get("ideal")),
                    "count": int(row.get("count", 0)),
                }

    particular_breakdown = []
    for p, pd_ in particular_scores.items():
        pct = round(pd_["score"] / pd_["ideal"] * 100, 1) if pd_["ideal"] > 0 else 0
        particular_breakdown.append({
            "particular": p[:80],
            "score":      round(pd_["score"], 1),
            "ideal":      round(pd_["ideal"], 1),
            "pct":        pct,
            "count":      pd_["count"],
        })
    particular_breakdown.sort(key=lambda x: x["pct"])

    # ── Misses list ───────────────────────────────────────────────────────────
    misses_list = []
    for sec_key, sec_label in SECTION_LABELS.items():
        for row in facet.get(f"{sec_key}_misses", []):
            ms = (row.get("misses") or "").strip()
            if not ms:
                continue
            sat = _parse_sat(row.get("submitted_at"))
            misses_list.append({
                "particular": (row.get("particular") or "").strip(),
                "section":    sec_label,
                "misses":     ms,
                "tl":         (row.get("tl")     or "").strip(),
                "client":     (row.get("client") or "").strip(),
                "date":       row.get("audit_date") or (sat.strftime("%d %b %Y") if sat else ""),
            })

    return {
        "section_breakdown":    section_breakdown,
        "particular_breakdown": particular_breakdown,
        "misses_list":          misses_list,
    }


# ─── export ───────────────────────────────────────────────────────────────────

@router.get("/export/{audit_id}")
async def export_audit(audit_id: str, current_user: str = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    email, is_admin, is_qa_only = _check_qa_access(current_user)
    try:
        oid = ObjectId(audit_id)
    except Exception:
        raise HTTPException(400, "Invalid ID")
    doc = qa_audit_collection.find_one({"_id": oid})
    if not doc:
        raise HTTPException(404, "Not found")
    if not (is_admin or is_qa_only) and doc.get("submitted_by") != current_user:
        raise HTTPException(403, "Access denied")

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"

    PURPLE     = "7B2D8B"; LIGHT_PURPLE = "E8D5F5"; BLUE_HDR = "1F4E79"
    ORANGE     = "C65911"; YELLOW       = "FFF2CC"; LIGHT_GREEN = "E2EFDA"
    LIGHT_GREY = "F2F2F2"; WHITE        = "FFFFFF"

    def cs(ws, row, col, value="", bold=False, bg=None, fc="000000",
           size=10, wrap=False, ha="left", va="center"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", bold=bold, size=size, color=fc)
        if bg: c.fill = PatternFill("solid", start_color=bg, end_color=bg)
        c.alignment = Alignment(wrap_text=wrap, horizontal=ha, vertical=va)
        thin = Side(style="thin", color="BBBBBB")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    def sec_hdr(ws, row, label, bg=BLUE_HDR, ncols=6):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill      = PatternFill("solid", start_color=bg, end_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    for col, w in [("A",46),("B",13),("C",32),("D",28),("E",9),("F",9)]:
        ws.column_dimensions[col].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    tc = ws.cell(row=r, column=1, value="JHS QUALITY AUDIT CHECKLIST")
    tc.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    tc.fill = PatternFill("solid", start_color=PURPLE, end_color=PURPLE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 28

    for label, val in [
        ("Project ID",             doc.get("project_id", "")),
        ("Project Name",           doc.get("project_name", "")),
        ("Client Name",            doc.get("client_name", "")),
        ("Type of Audit",          doc.get("type_of_audit", "")),
        ("Audit Period",           f"From: {doc.get('audit_period_from','')}   To: {doc.get('audit_period_to','')}"),
        ("Quality Audit Given By", f"{doc.get('audit_given_by_name','')}  (Emp ID: {doc.get('audit_given_by_emp_id','')})"),
        ("Project TL / Manager",   doc.get("project_tl", "")),
        ("Project PnD",            doc.get("project_pnd", "")),
        ("Audit Date",             doc.get("audit_date", "")),
    ]:
        r += 1
        ws.row_dimensions[r].height = 15
        cs(ws, r, 1, label, bold=True, bg=LIGHT_PURPLE, fc=PURPLE)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cs(ws, r, 2, val, bg=WHITE)

    r += 1
    ws.row_dimensions[r].height = 20
    for ci, h in enumerate(["Particulars","YES/NO/NA","Evidence / Remark","Misses / Serious","Score","Ideal"], 1):
        cs(ws, r, ci, h, bold=True, bg=BLUE_HDR, fc=WHITE, ha="center")

    for sec_key, sec_label, sec_color in [
        ("section_a", "(A) Audit Planning",  BLUE_HDR),
        ("section_b", "(B) Audit Execution", "375623"),
        ("section_c", "(C) Audit Reporting", ORANGE),
        ("section_d", "(D) Quality Control", PURPLE),
        ("section_e", "(E) Optimization",    "1E6B3C"),
    ]:
        r += 1
        sec_hdr(ws, r, sec_label, bg=sec_color)
        for item in doc.get(sec_key, []):
            r += 1
            ws.row_dimensions[r].height = 28
            resp = item.get("response", "")
            rb   = "E2EFDA" if resp == "YES" else "FFCCCC" if resp == "NO" else LIGHT_GREY
            cs(ws, r, 1, item.get("particular", ""),  bg=LIGHT_GREY, wrap=True,   va="top")
            cs(ws, r, 2, resp,                         bg=rb,          ha="center", bold=True)
            cs(ws, r, 3, item.get("evidence", ""),     bg=WHITE,       wrap=True,  va="top")
            cs(ws, r, 4, item.get("misses", ""),       bg="FFF8E1",    wrap=True,  va="top")
            cs(ws, r, 5, item.get("score", ""),        bg=YELLOW,      ha="center", bold=True)
            cs(ws, r, 6, item.get("ideal_score", ""),  bg=LIGHT_GREEN, ha="center")

    improvements = (doc.get("improvements") or "").strip()
    if improvements:
        r += 1
        sec_hdr(ws, r, "Improvements (if any)", bg="1565C0")
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=6)
        c = ws.cell(row=r, column=1, value=improvements)
        c.font      = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.fill      = PatternFill("solid", start_color="E3F2FD", end_color="E3F2FD")
        ws.row_dimensions[r].height = 60
        r += 4

    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cs(ws, r, 1, "TOTAL",                     bold=True, bg=BLUE_HDR,    fc=WHITE,  ha="right")
    cs(ws, r, 5, _safe_float(doc.get("total_score")), bold=True, bg=YELLOW,      ha="center")
    cs(ws, r, 6, _safe_float(doc.get("ideal_score")), bold=True, bg=LIGHT_GREEN, ha="center")
    r += 1
    ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    sc  = _safe_float(doc.get("total_score"))
    isc = _safe_float(doc.get("ideal_score"))
    sca = round(sc * 100 / isc, 2) if isc > 0 else _safe_float(doc.get("scaled_total"))
    cs(ws, r, 1, "SCALED TOTAL (out of 100)", bold=True, bg=PURPLE,      fc=WHITE,  ha="right")
    cs(ws, r, 5, sca,                          bold=True, bg=LIGHT_PURPLE, ha="center", fc=PURPLE)
    cs(ws, r, 6, 100,                          bold=True, bg=LIGHT_GREEN, ha="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe  = (doc.get("client_name") or "Audit").replace(" ", "_")[:30]
    dstr  = (doc.get("audit_date")  or "")[:10].replace("/", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="QA_{safe}_{dstr}.xlsx"'},
    )